import openpyxl
import crawl
def list_sheet_names(filepath):
    wb = openpyxl.load_workbook(filepath, read_only=True)
    return wb.sheetnames


if __name__ == "__main__":
    import os

    root_dir = os.path.dirname(os.path.abspath(__file__))
    excel_files = []
    for dirpath, _, filenames in os.walk(root_dir):
        for f in filenames:
            if f.endswith('.xlsx'):
                full_path = os.path.join(dirpath, f)
                excel_files.append(full_path)

    if not excel_files:
        print("No Excel (.xlsx) files found in the current directory.")
    else:
        print("Available Excel files:")
        for i, file in enumerate(excel_files, 1):
            print(f"{i}. {file}")
        
        choice = input("Enter the number of the Excel file to read: ")

        try:
            selected_file = excel_files[int(choice) - 1]
            sheet_names = list_sheet_names(selected_file)
            print("Sheet names:")
            for name in sheet_names:
                print(f"- {name}")
            
            # Load workbook and read column B2 to B317
            wb = openpyxl.load_workbook(selected_file, read_only=True)
            ws = wb.active  # Use the first sheet by default
            cell_values = [ws[f"B{i}"].value for i in range(2, 318) if ws[f"B{i}"].value]

            import subprocess
            subprocess.run(["python", "crawl.py"] + cell_values)
        except (IndexError, ValueError):
            print("Invalid selection.")
