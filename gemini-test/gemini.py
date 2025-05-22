import os
import json
from dotenv import load_dotenv
from openpyxl import load_workbook
import google.generativeai as genai

# Load environment variables
load_dotenv()
api_key = os.getenv("GEMINI_API_KEY")
if not api_key:
    print("Error: GEMINI_API_KEY is not set. Please check your .env file.")
    exit(1)
genai.configure(api_key=api_key)

def read_names_from_excel(file_path, start_row=2, end_row=None):
    wb = load_workbook(filename=file_path)
    ws = wb.active
    names = []
    row = start_row
    while True:
        cell = ws[f'B{row}']
        if cell.value is None:
            break
        if end_row and row > end_row:
            break
        names.append(cell.value.strip())
        row += 1
    return names

def find_images_with_gemini(names):
    model = genai.GenerativeModel(model_name="gemini-1.5-flash")
    import time
    for name in names:
        prompt = (
            f"Generate a list of search keywords for Google Images or Bing to find the most relevant and appropriate pictures "
            f"representing the concept: \"{name}\". Include geographic or regional context when applicable (e.g. country, region, or local terminology). "
            f"Format the output as plain text, one search phrase per line. Each phrase should be suitable for directly pasting into an image search engine."
        )
        response = model.generate_content(prompt)
        print(f"\nResults for '{name}':\n{response.text}\n")
        keywords = response.text.strip().split('\n')
        keywords = [kw.strip('-• ').strip() for kw in keywords if kw.strip()]
        safe_name = "".join(c if c.isalnum() or c in (' ', '_', '-') else '_' for c in name)
        file_name = f"{safe_name.replace(' ', '_')}.json"
        os.makedirs(os.path.dirname(file_name) or ".", exist_ok=True)
        with open(file_name, 'w') as f:
            json.dump({"keywords": keywords}, f, indent=2)
        print(f"Saved keywords to {file_name}")
        time.sleep(2)

if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="Search Gemini for images based on names in Excel file.")
    parser.add_argument("excel_file", help="Path to the Excel file.")
    parser.add_argument("--start", type=int, default=2, help="Start row (default is 2, skipping header).")
    parser.add_argument("--end", type=int, help="End row (optional).")
    parser.add_argument("--all", action="store_true", help="Read all values until an empty cell.")
    parser.add_argument("--batch", action="store_true", help="Run the search on all names without prompting.")

    args = parser.parse_args()

    if args.all:
        names = read_names_from_excel(args.excel_file, start_row=args.start)
    else:
        names = read_names_from_excel(args.excel_file, start_row=args.start, end_row=args.end)

    if args.batch:
        find_images_with_gemini(names)
    else:
        print("Select a name to search for images:")
        print("0. Run search for all names")
        for i, name in enumerate(names, 1):
            print(f"{i}. {name}")
        choice = int(input("Enter the number of the name you want to search for: "))
        if choice == 0:
            find_images_with_gemini(names)
        else:
            selected_name = names[choice - 1]
            find_images_with_gemini([selected_name])